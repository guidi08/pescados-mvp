#!/usr/bin/env python3
import json
import os
import re
import subprocess
from datetime import datetime
from pathlib import Path

import pandas as pd

ACCOUNT = "asanamomo1@gmail.com"
OAUTH_CLIENT = "asana"
OUT_DIR = "/Users/guilhermesbot/clawd/anexos/momo_reports"
STATE_PATH = "/Users/guilhermesbot/clawd/automation/auto_momo_report_state.json"
WHATSAPP_TO = "+5511999713995"

Path(OUT_DIR).mkdir(parents=True, exist_ok=True)


def sh(cmd):
    env = os.environ.copy()
    env["PATH"] = "/opt/homebrew/bin:/usr/local/bin:/usr/bin:/bin:/usr/sbin:/sbin"
    return subprocess.check_output(cmd, shell=True, text=True, env=env)


def load_state():
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH, "r") as f:
            return json.load(f)
    return {"processed_message_ids": [], "last_vouchers_total": 37463.92, "flow_base_leo": 49595.07, "flow_base_pin": 53995.08}


def save_state(state):
    with open(STATE_PATH, "w") as f:
        json.dump(state, f, indent=2)


def safe_name(name: str) -> str:
    name = name.strip() or "arquivo"
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
    return name[:180]


def fmt_money(v):
    if v is None or pd.isna(v):
        return "—"
    return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_int(v):
    if v is None or pd.isna(v):
        return "—"
    return str(int(v))


def parse_export(path):
    df = pd.read_excel(path, sheet_name="Export_CSV", header=1)
    df = df.dropna(how="all")
    if df.empty:
        return None

    def to_date(val):
        if pd.isna(val):
            return None
        try:
            return pd.to_datetime(val, unit="D", origin="1899-12-30").date()
        except Exception:
            return val

    if len(df) == 1:
        row = df.iloc[0]

        pag_leop = row.get("total_pagamentos")
        pag_pin = 0
        entradas_pin = 0
        # try to read payments + entradas Pinheiros from sheet "1 dias"
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            if "1 dias" in wb.sheetnames:
                ws = wb["1 dias"]
                for r in range(1, ws.max_row + 1):
                    label = ws.cell(r, 1).value
                    if label == "Total pagamentos Leopoldina(R$)":
                        val = ws.cell(r, 2).value
                        if val is not None:
                            pag_leop = val
                    if label == "Total pagamentos Pinheiros(R$)":
                        val = ws.cell(r, 2).value
                        if val is not None:
                            pag_pin = val
                    if isinstance(label, str) and label.lower() in ("keeta pinheiros", "99 pinheiros", "ifood pinheiros"):
                        val = ws.cell(r, 2).value
                        if val is not None:
                            entradas_pin += float(val)
            wb.close()
        except Exception:
            pass

        return {
            "is_multi": False,
            "date": to_date(row.get("data_movimento")),
            "total_entradas": row.get("total_entradas"),
            "total_pagamentos": row.get("total_pagamentos"),
            "salao_total": row.get("salao_total"),
            "delivery_leop_total_valor": row.get("delivery_leop_total_valor"),
            "delivery_leop_total_qtd": row.get("delivery_leop_total_qtd"),
            "delivery_pin_total_valor": row.get("delivery_pin_total_valor"),
            "delivery_pin_total_qtd": row.get("delivery_pin_total_qtd"),
            "rod_total_dia": row.get("rod_total_dia"),
            "vouchers_total": row.get("vouchers_total"),
            "pag_leop": pag_leop,
            "pag_pin": pag_pin,
            "entradas_pin": entradas_pin,
        }

    # multi-day (weekend) -> sum totals, keep last voucher as current
    start_date = to_date(df.iloc[0].get("data_movimento"))
    end_date = to_date(df.iloc[-1].get("data_movimento"))

    def col_sum(name):
        return df[name].sum(skipna=True) if name in df.columns else None

    vouchers_total = df["vouchers_total"].dropna().iloc[-1] if "vouchers_total" in df.columns and not df["vouchers_total"].dropna().empty else None

    return {
        "is_multi": True,
        "start_date": start_date,
        "end_date": end_date,
        "total_entradas": col_sum("total_entradas"),
        "total_pagamentos": col_sum("total_pagamentos"),
        "salao_total": col_sum("salao_total"),
        "delivery_leop_total_valor": col_sum("delivery_leop_total_valor"),
        "delivery_leop_total_qtd": col_sum("delivery_leop_total_qtd"),
        "delivery_pin_total_valor": col_sum("delivery_pin_total_valor"),
        "delivery_pin_total_qtd": col_sum("delivery_pin_total_qtd"),
        "rod_total_dia": col_sum("rod_total_dia"),
        "vouchers_total": vouchers_total,
        "pag_leop": total_pagamentos,
        "pag_pin": 0,
    }


def summarize(data, last_vouchers_total, flow_base_leo, flow_base_pin):
    if data.get("is_multi"):
        # for multi-day reports, do not show detailed per-unit payments
        data["pag_leop"] = data.get("total_pagamentos")
        data["pag_pin"] = 0
        start = data.get("start_date")
        end = data.get("end_date")
        if hasattr(start, "strftime") and hasattr(end, "strftime"):
            date_str = f"Fim de semana {start.strftime('%d/%m')}–{end.strftime('%d/%m/%Y')}"
        else:
            date_str = "Fim de semana"
    else:
        date = data.get("date")
        if isinstance(date, datetime):
            date = date.date()
        date_str = date.strftime("%d/%m/%Y") if hasattr(date, "strftime") else str(date)

    vouchers_total = data.get("vouchers_total")
    if vouchers_total is not None and last_vouchers_total is not None:
        voucher_delta = float(vouchers_total) - float(last_vouchers_total)
    else:
        voucher_delta = None

    total_entradas = data.get("total_entradas")
    if voucher_delta is not None and total_entradas is not None:
        soma_voucher_entradas = float(voucher_delta) + float(total_entradas)
    else:
        soma_voucher_entradas = None

    total_pagamentos = data.get("total_pagamentos")
    pag_leop = data.get("pag_leop")
    pag_pin = data.get("pag_pin")
    entradas_pin = data.get("entradas_pin", 0)

    # fluxo por unidade
    if soma_voucher_entradas is not None and pag_leop is not None:
        fluxo_gerado_leo = float(soma_voucher_entradas) - float(pag_leop)
    else:
        fluxo_gerado_leo = None

    if entradas_pin is not None and pag_pin is not None:
        fluxo_gerado_pin = float(entradas_pin) - float(pag_pin)
    else:
        fluxo_gerado_pin = None

    # fluxo total (geral)
    if total_pagamentos is not None and soma_voucher_entradas is not None:
        fluxo_gerado = float(soma_voucher_entradas) - float(total_pagamentos)
    else:
        fluxo_gerado = None

    # Fluxo de caixa: Leopoldina usa o fluxo de ontem como saldo inicial de hoje
    if flow_base_leo is not None:
        fluxo_inicio_leo = float(flow_base_leo)
    else:
        fluxo_inicio_leo = None

    if flow_base_pin is not None:
        fluxo_inicio_pin = float(flow_base_pin)
    else:
        fluxo_inicio_pin = None

    if fluxo_gerado_leo is not None and fluxo_inicio_leo is not None:
        fluxo_pos_pagamentos_leo = float(fluxo_inicio_leo) + float(fluxo_gerado_leo)
    else:
        fluxo_pos_pagamentos_leo = None

    if fluxo_gerado_pin is not None and fluxo_inicio_pin is not None:
        fluxo_pos_pagamentos_pin = float(fluxo_inicio_pin) + float(fluxo_gerado_pin)
    else:
        fluxo_pos_pagamentos_pin = None

    total_pedidos = (data.get("delivery_leop_total_qtd") or 0) + (data.get("delivery_pin_total_qtd") or 0)

    total_vendas = None
    if data.get('salao_total') is not None and data.get('delivery_leop_total_valor') is not None and data.get('delivery_pin_total_valor') is not None:
        total_vendas = float(data.get('salao_total')) + float(data.get('delivery_leop_total_valor')) + float(data.get('delivery_pin_total_valor'))

    lines = [
        f"Relatório {date_str}",
        f"• Total vendas (salão + Leo + Pin): {fmt_money(total_vendas)}",
        f"• Pedidos total: {fmt_int(total_pedidos)} (Leo: {fmt_int(data.get('delivery_leop_total_qtd'))} | Pin: {fmt_int(data.get('delivery_pin_total_qtd'))})",
        f"• Entrada financeira no dia (total): {fmt_money(total_entradas)}",
        f"• Pagamentos (saídas) no dia (total): {fmt_money(total_pagamentos)}",
        f"• Voucher total: {fmt_money(vouchers_total)}",
        f"• Entrada no voucher (Δ): {fmt_money(voucher_delta)}",
        f"• Δ voucher + entrada: {fmt_money(soma_voucher_entradas)}",
        "",
        "LEOPOLDINA",
        f"• Vendas salão (dia anterior): {fmt_money(data.get('salao_total'))}",
        f"• Delivery Leopoldina: {fmt_money(data.get('delivery_leop_total_valor'))}",
        f"• Rodízios vendidos: {fmt_int(data.get('rod_total_dia'))}",
        f"• Pedidos Leopoldina: {fmt_int(data.get('delivery_leop_total_qtd'))}",
        f"• Pagamentos Leopoldina: {fmt_money(data.get('pag_leop'))}",
        f"• Fluxo gerado no dia (Leopoldina): {fmt_money(fluxo_gerado_leo)}",
        f"• Fluxo de caixa no início do dia (Leopoldina): {fmt_money(fluxo_inicio_leo)}",
        f"• Fluxo de caixa após pagamentos (Leopoldina): {fmt_money(fluxo_pos_pagamentos_leo)}",
        "",
        "PINHEIROS",
        f"• Delivery Pinheiros: {fmt_money(data.get('delivery_pin_total_valor'))}",
        f"• Pedidos Pinheiros: {fmt_int(data.get('delivery_pin_total_qtd'))}",
        f"• Pagamentos Pinheiros: {fmt_money(data.get('pag_pin'))}",
        f"• Fluxo gerado no dia (Pinheiros): {fmt_money(fluxo_gerado_pin)}",
        f"• Fluxo de caixa no início do dia (Pinheiros): {fmt_money(fluxo_inicio_pin)}",
        f"• Fluxo de caixa após pagamentos (Pinheiros): {fmt_money(fluxo_pos_pagamentos_pin)}",
    ]
    return "\n".join(lines)


def main():
    state = load_state()
    processed = set(state.get("processed_message_ids", []))
    last_vouchers_total = state.get("last_vouchers_total", 37463.92)
    flow_base_leo = state.get("flow_base_leo", 49595.07)
    flow_base_pin = state.get("flow_base_pin", 53995.08)

    msg_json = sh(f"gog gmail messages search \"in:inbox newer_than:7d has:attachment\" --account {ACCOUNT} --client {OAUTH_CLIENT} --max 25 --json")
    msg = json.loads(msg_json)
    messages = msg.get("messages", [])

    for m in messages:
        mid = m.get("id")
        if not mid or mid in processed:
            continue

        full = json.loads(sh(f"gog gmail get {mid} --account {ACCOUNT} --client {OAUTH_CLIENT} --json"))
        attachments = full.get("attachments", [])
        if not attachments:
            processed.add(mid)
            continue

        headers = full.get("headers", {})
        subject = headers.get("subject", "")

        for att in attachments:
            filename = att.get("filename") or "arquivo"
            if not filename.lower().endswith(".xlsx"):
                continue
            att_id = att.get("attachmentId")
            if not att_id:
                continue

            base = safe_name(filename)
            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
            out_name = f"{ts}_{mid}_{base}"
            out_path = os.path.join(OUT_DIR, out_name)

            sh(f"gog gmail attachment {mid} {att_id} --account {ACCOUNT} --client {OAUTH_CLIENT} --out \"{out_path}\"")

            data = parse_export(out_path)
            if data:
                prev_vouchers_total = last_vouchers_total
                msg_text = summarize(data, last_vouchers_total, flow_base_leo, flow_base_pin)
                sh(f"clawdbot message send --channel whatsapp --target {WHATSAPP_TO} --message \"{msg_text}\"")

                # update flow base for next day: flow_base + fluxo_gerado
                try:
                    vouchers_total = data.get("vouchers_total")
                    total_entradas = data.get("total_entradas")
                    if vouchers_total is not None and total_entradas is not None:
                        voucher_delta = float(vouchers_total) - float(prev_vouchers_total)
                        soma_voucher_entradas = float(voucher_delta) + float(total_entradas)
                        pag_leop = data.get("pag_leop") or 0
                        pag_pin = data.get("pag_pin") or 0
                        fluxo_gerado_leo = float(soma_voucher_entradas) - float(pag_leop)
                        fluxo_gerado_pin = float(data.get("entradas_pin", 0)) - float(pag_pin)
                        fluxo_pos_pagamentos_leo = float(flow_base_leo) + float(fluxo_gerado_leo)
                        fluxo_pos_pagamentos_pin = float(flow_base_pin) + float(fluxo_gerado_pin)
                        flow_base_leo = float(fluxo_pos_pagamentos_leo)
                        flow_base_pin = float(fluxo_pos_pagamentos_pin)
                except Exception:
                    pass

                # update last vouchers total for next report
                if data.get("vouchers_total") is not None:
                    last_vouchers_total = float(data.get("vouchers_total"))
            else:
                msg_text = f"Relatório recebido, mas não consegui ler o Export_CSV: {base}\nAssunto: {subject}"
                sh(f"clawdbot message send --channel whatsapp --target {WHATSAPP_TO} --message \"{msg_text}\"")

        processed.add(mid)

    state["processed_message_ids"] = sorted(list(processed))
    state["last_vouchers_total"] = last_vouchers_total
    state["flow_base_leo"] = flow_base_leo
    state["flow_base_pin"] = flow_base_pin
    save_state(state)


if __name__ == "__main__":
    main()
